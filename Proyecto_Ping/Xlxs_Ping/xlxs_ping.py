from openpyxl import Workbook, load_workbook
import os
from Variables_Ping.variables_ping import filas, columnas, ruta, fecha

####################__Excel__####################
def reg_sale():
    if os.path.exists(f"{ruta}Registro {fecha}.xlsx"):
        wb = load_workbook(filename = f"{ruta}Registro {fecha}.xlsx")
        ws = wb.active
        for f in filas:
            ws.append(f)
        wb.save(f"{ruta}Registro {fecha}.xlsx")
    else:
        wb = Workbook()
        ws = wb.active
        for c in columnas:
            ws.append(c)
        for f in filas:
            ws.append(f)
        wb.save(f"{ruta}Registro {fecha}.xlsx")